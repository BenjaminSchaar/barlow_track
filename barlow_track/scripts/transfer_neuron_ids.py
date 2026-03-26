#!/usr/bin/env python
"""
Transfer neuron IDs from annotated source recordings to an unannotated target recording
using a trained cross-recording Barlow model and consensus voting.

Workflow:
  1. Embed source recordings (annotated) with model.backbone()
  2. Embed target recording (unannotated)
  3. For each target neuron, find best match in each source → vote
  4. Consensus: if N/M sources agree on same ID → accept

Usage:
    python transfer_neuron_ids.py \
        --model /path/to/resnet50.pth \
        --sources /path/to/source_rec_1 /path/to/source_rec_2 /path/to/source_rec_3 \
        --target /path/to/target_recording \
        --consensus_threshold 2 \
        [--n_frames 500] \
        [--output /custom/output/dir]

Output:
    target_recording/3-tracking/id_transfer/
        candidate_matches.csv     — all matches with consensus scores
        transfer_summary.json     — summary statistics
        transfer_config.yaml      — reproducibility: model, sources, params
"""

import argparse
import csv
import json
import logging
import pickle
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import torch
from tqdm.auto import tqdm

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(name)s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)


def load_annotation_names(project_path: str) -> dict:
    """
    Load neuron ID1 names from manual_annotation.xlsx.

    Returns dict: neuron_tracking_id -> biological_name (only for annotated neurons).
    """
    from openpyxl import load_workbook

    anno_path = Path(project_path) / "3-tracking" / "manual_annotation" / "manual_annotation.xlsx"
    if not anno_path.exists():
        logger.warning(f"No manual_annotation.xlsx found at {anno_path}")
        return {}

    wb = load_workbook(str(anno_path), read_only=True)
    ws = wb.active

    name_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 3 and row[0] is not None and row[2] is not None:
            neuron_id = str(row[0]).strip()
            bio_name = str(row[2]).strip()  # Column C = ID1
            if bio_name and bio_name.lower() not in ("none", ""):
                name_map[neuron_id] = bio_name
    wb.close()

    return name_map


def embed_recording(
    project_path: str,
    model,
    device: torch.device,
    target_sz: np.ndarray,
    n_sample_frames: int = 500,
) -> dict:
    """
    Embed one recording using model.backbone().

    Returns dict: neuron_tracking_id -> mean_embedding (np.ndarray of shape (D,))
    Also stores per-neuron frame count for diagnostics.
    """
    from wbfm.utils.projects.finished_project_data import ProjectData
    from barlow_track.utils.data_loading import get_bbox_data_for_volume_with_label

    project_data = ProjectData.load_final_project_data(
        project_path, allow_hybrid_loading=True
    )

    max_frames = project_data.num_frames
    n_sample = min(n_sample_frames, max_frames)
    frame_indices = np.linspace(0, max_frames - 1, n_sample, dtype=int)

    neuron_embeddings = defaultdict(list)

    for t in tqdm(frame_indices, desc=f"Embedding {Path(project_path).name}", leave=False):
        try:
            all_dat_dict, seg2name, _ = get_bbox_data_for_volume_with_label(
                project_data, int(t), target_sz=target_sz, include_untracked=True
            )
        except (KeyError, IndexError, FileNotFoundError):
            continue

        if not all_dat_dict:
            continue

        names = list(all_dat_dict.keys())
        crops = np.stack([all_dat_dict[n] for n in names], 0).astype(np.float32)
        crops_tensor = torch.from_numpy(crops).unsqueeze(1).to(device)

        with torch.no_grad():
            embeddings = model.backbone(crops_tensor).cpu().numpy()

        for name, emb in zip(names, embeddings):
            neuron_embeddings[name].append(emb)

    # Compute mean embedding per neuron
    mean_embeddings = {}
    frame_counts = {}
    for neuron_id, emb_list in neuron_embeddings.items():
        mean_embeddings[neuron_id] = np.mean(emb_list, axis=0)
        frame_counts[neuron_id] = len(emb_list)

    logger.info(f"  {Path(project_path).name}: {len(mean_embeddings)} neurons embedded "
                f"(mean {np.mean(list(frame_counts.values())):.0f} frames/neuron)")
    return mean_embeddings, frame_counts


def cosine_similarity_matrix(embeddings_a: np.ndarray, embeddings_b: np.ndarray) -> np.ndarray:
    """Compute cosine similarity matrix between two sets of embeddings."""
    # Normalize
    norms_a = np.linalg.norm(embeddings_a, axis=1, keepdims=True)
    norms_b = np.linalg.norm(embeddings_b, axis=1, keepdims=True)
    norms_a = np.maximum(norms_a, 1e-8)
    norms_b = np.maximum(norms_b, 1e-8)
    a_normed = embeddings_a / norms_a
    b_normed = embeddings_b / norms_b
    return a_normed @ b_normed.T


def consensus_matching(
    target_embeddings: dict,
    source_data: list,
    consensus_threshold: int,
) -> list:
    """
    For each target neuron, find best match in each source recording,
    then apply consensus voting.

    Parameters
    ----------
    target_embeddings : dict
        neuron_id -> mean_embedding for target recording
    source_data : list of dicts, each with:
        'embeddings': {neuron_id: mean_embedding}
        'annotations': {neuron_id: biological_name}
        'recording_name': str
    consensus_threshold : int
        Minimum number of sources that must agree for a match to be accepted

    Returns
    -------
    list of dicts, one per target neuron, sorted by consensus descending
    """
    target_ids = sorted(target_embeddings.keys())
    target_emb_matrix = np.stack([target_embeddings[nid] for nid in target_ids], 0)

    n_sources = len(source_data)
    results = []

    for src_idx, src in enumerate(source_data):
        # Only use annotated source neurons
        annotated_ids = [nid for nid in src["embeddings"] if nid in src["annotations"]]
        if not annotated_ids:
            logger.warning(f"Source {src['recording_name']} has no annotated neurons, skipping")
            continue

        src_emb_matrix = np.stack([src["embeddings"][nid] for nid in annotated_ids], 0)
        src_bio_names = [src["annotations"][nid] for nid in annotated_ids]

        sim = cosine_similarity_matrix(target_emb_matrix, src_emb_matrix)

        # For each target neuron, find best match in this source
        best_idx = np.argmax(sim, axis=1)
        best_sim = np.max(sim, axis=1)

        for i, target_id in enumerate(target_ids):
            if src_idx == 0:
                results.append({
                    "target_neuron_id": target_id,
                    "votes": [],
                    "similarities": [],
                    "source_recordings": [],
                })
            results[i]["votes"].append(src_bio_names[best_idx[i]])
            results[i]["similarities"].append(float(best_sim[i]))
            results[i]["source_recordings"].append(src["recording_name"])

    # Compute consensus for each target neuron
    output = []
    for r in results:
        vote_counts = Counter(r["votes"])
        if not vote_counts:
            continue

        top_name, top_count = vote_counts.most_common(1)[0]
        consensus = top_count / n_sources
        mean_sim_for_top = np.mean([
            s for v, s in zip(r["votes"], r["similarities"]) if v == top_name
        ])

        # Per-source vote detail
        vote_detail = "; ".join(
            f"{rec}={vote}({sim:.3f})"
            for rec, vote, sim in zip(r["source_recordings"], r["votes"], r["similarities"])
        )

        accepted = top_count >= consensus_threshold
        output.append({
            "target_neuron_id": r["target_neuron_id"],
            "predicted_id": top_name,
            "consensus_count": top_count,
            "consensus_fraction": round(consensus, 3),
            "n_sources": n_sources,
            "mean_similarity": round(mean_sim_for_top, 4),
            "accepted": "yes" if accepted else "no",
            "vote_detail": vote_detail,
        })

    # Sort: accepted first, then by consensus descending, then by similarity
    output.sort(key=lambda x: (-int(x["accepted"] == "yes"), -x["consensus_count"], -x["mean_similarity"]))
    return output


def load_model(model_path: str, device: torch.device):
    """Load trained Barlow model and return (model, target_sz)."""
    from barlow_track.utils.barlow import BarlowTwins3d
    from barlow_track.utils.siamese import ResidualEncoder3D

    args_path = Path(model_path).parent / "args.pickle"
    with open(args_path, "rb") as f:
        model_args = pickle.load(f)

    target_sz = np.array([model_args.target_sz_z, model_args.target_sz_xy, model_args.target_sz_xy])

    try:
        user_args = vars(vars(model_args).get("backbone_kwargs", dict()))
    except TypeError:
        user_args = dict()

    backbone_kwargs = dict(
        in_channels=1,
        num_levels=user_args.get("num_levels", 2),
        f_maps=user_args.get("f_maps", 4),
        crop_sz=target_sz,
    )
    model = BarlowTwins3d(model_args, backbone=ResidualEncoder3D, **backbone_kwargs).to(device)
    model.load_state_dict(torch.load(model_path, map_location=device))
    model.eval()

    return model, target_sz


def main():
    parser = argparse.ArgumentParser(
        description="Transfer neuron IDs via consensus matching",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--model", required=True, help="Path to trained resnet50.pth")
    parser.add_argument("--sources", required=True, nargs="+",
                        help="Paths to annotated source recordings")
    parser.add_argument("--target", required=True, help="Path to target recording")
    parser.add_argument("--consensus_threshold", type=int, default=2,
                        help="Min sources that must agree (default: 2)")
    parser.add_argument("--n_frames", type=int, default=500,
                        help="Frames to sample per recording (default: 500)")
    parser.add_argument("--output", default=None,
                        help="Output directory (default: target/3-tracking/id_transfer/)")
    args = parser.parse_args()

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    # Output directory
    if args.output:
        output_dir = Path(args.output)
    else:
        output_dir = Path(args.target) / "3-tracking" / "id_transfer"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Validate
    n_sources = len(args.sources)
    if args.consensus_threshold > n_sources:
        logger.error(f"consensus_threshold ({args.consensus_threshold}) > n_sources ({n_sources})")
        return
    if args.consensus_threshold < 1:
        logger.error("consensus_threshold must be >= 1")
        return

    logger.info(f"Model: {args.model}")
    logger.info(f"Sources: {n_sources} recordings")
    logger.info(f"Target: {args.target}")
    logger.info(f"Consensus threshold: {args.consensus_threshold}/{n_sources}")

    # Load model
    logger.info("Loading model...")
    model, target_sz = load_model(args.model, device)
    logger.info(f"Model loaded (target_sz={target_sz})")

    # Embed source recordings
    source_data = []
    for src_path in args.sources:
        logger.info(f"=== Embedding source: {Path(src_path).name} ===")
        annotations = load_annotation_names(src_path)
        embeddings, frame_counts = embed_recording(
            src_path, model, device, target_sz, args.n_frames
        )
        n_annotated = sum(1 for nid in embeddings if nid in annotations)
        logger.info(f"  {n_annotated}/{len(embeddings)} neurons have annotations")

        source_data.append({
            "embeddings": embeddings,
            "annotations": annotations,
            "recording_name": Path(src_path).name,
            "frame_counts": frame_counts,
        })

    # Check annotation overlap across sources
    all_bio_names = [set(s["annotations"].values()) for s in source_data]
    shared_across_all = set.intersection(*all_bio_names) if all_bio_names else set()
    logger.info(f"Neuron types shared across all {n_sources} sources: {len(shared_across_all)}")

    # Embed target
    logger.info(f"=== Embedding target: {Path(args.target).name} ===")
    target_annotations = load_annotation_names(args.target)
    target_embeddings, target_frame_counts = embed_recording(
        args.target, model, device, target_sz, args.n_frames
    )
    n_target_annotated = sum(1 for nid in target_embeddings if nid in target_annotations)
    logger.info(f"  Target has {n_target_annotated}/{len(target_embeddings)} existing annotations "
                f"(will be used for accuracy evaluation)")

    # Consensus matching
    logger.info("=== Running consensus matching ===")
    matches = consensus_matching(
        target_embeddings, source_data, args.consensus_threshold
    )

    # Evaluate against existing annotations (if target has any)
    n_accepted = sum(1 for m in matches if m["accepted"] == "yes")
    n_correct = 0
    n_evaluable = 0
    for m in matches:
        if m["accepted"] == "yes" and m["target_neuron_id"] in target_annotations:
            n_evaluable += 1
            if m["predicted_id"] == target_annotations[m["target_neuron_id"]]:
                n_correct += 1

    if n_evaluable > 0:
        accuracy = n_correct / n_evaluable
        logger.info(f"Accuracy on already-annotated neurons: {accuracy:.1%} ({n_correct}/{n_evaluable})")
    else:
        accuracy = None
        logger.info("No existing annotations in target to evaluate against")

    # Save candidate_matches.csv
    csv_path = output_dir / "candidate_matches.csv"
    if matches:
        fieldnames = list(matches[0].keys())
        with open(csv_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(matches)
    logger.info(f"Saved {len(matches)} matches to {csv_path}")

    # Save transfer_summary.json
    summary = {
        "model_path": str(args.model),
        "source_recordings": [str(p) for p in args.sources],
        "target_recording": str(args.target),
        "n_sources": n_sources,
        "consensus_threshold": args.consensus_threshold,
        "n_frames": args.n_frames,
        "n_target_neurons": len(target_embeddings),
        "n_accepted": n_accepted,
        "n_total_matches": len(matches),
        "n_shared_types_across_sources": len(shared_across_all),
        "accuracy_on_existing_annotations": accuracy,
        "n_evaluable": n_evaluable,
        "n_correct": n_correct,
    }
    with open(output_dir / "transfer_summary.json", "w") as f:
        json.dump(summary, f, indent=2)

    # Save transfer_config.yaml for reproducibility
    config = {
        "model_path": str(args.model),
        "source_recordings": [str(p) for p in args.sources],
        "target_recording": str(args.target),
        "consensus_threshold": args.consensus_threshold,
        "n_frames": args.n_frames,
    }
    from ruamel.yaml import YAML
    yaml_writer = YAML()
    with open(output_dir / "transfer_config.yaml", "w") as f:
        yaml_writer.dump(config, f)

    # Print summary
    print(f"\n{'='*70}")
    print(f"TRANSFER SUMMARY")
    print(f"{'='*70}")
    print(f"Source recordings:     {n_sources}")
    print(f"Consensus threshold:   {args.consensus_threshold}/{n_sources}")
    print(f"Target neurons:        {len(target_embeddings)}")
    print(f"Accepted matches:      {n_accepted}")
    if accuracy is not None:
        print(f"Accuracy (existing):   {accuracy:.1%} ({n_correct}/{n_evaluable})")
    print(f"")

    # Consensus breakdown
    consensus_counts = Counter(m["consensus_count"] for m in matches)
    print(f"Consensus breakdown:")
    for count in sorted(consensus_counts.keys(), reverse=True):
        n = consensus_counts[count]
        marker = " <-- threshold" if count == args.consensus_threshold else ""
        accepted_marker = " (accepted)" if count >= args.consensus_threshold else ""
        print(f"  {count}/{n_sources} sources agree: {n:>4} neurons{accepted_marker}{marker}")

    print(f"\nResults saved to: {output_dir}")
    print(f"  candidate_matches.csv   — review matches before applying")
    print(f"  transfer_summary.json   — summary statistics")
    print(f"  transfer_config.yaml    — reproducibility config")
    print(f"\nNext step:")
    print(f"  python apply_transfer.py --target {args.target} --matches {csv_path}")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()
