#!/usr/bin/env python
"""
Apply transferred neuron IDs to a target recording's manual_annotation.xlsx.

Reads candidate_matches.csv from transfer_neuron_ids.py and writes accepted
biological names into the target's ID1 column (column C).

Creates a timestamped backup of the original file before writing.

Usage:
    python apply_transfer.py \
        --target /path/to/target_recording \
        --matches /path/to/candidate_matches.csv \
        [--min_consensus 0.66] \
        [--overwrite_existing]

Safety:
    - Backs up manual_annotation.xlsx before any modification
    - By default, does NOT overwrite neurons that already have an ID1 value
    - Use --overwrite_existing to replace existing annotations
    - Dry-run mode (--dry_run) shows what would change without writing
"""

import argparse
import csv
import json
import logging
import shutil
from datetime import datetime
from pathlib import Path

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(name)s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)


def main():
    parser = argparse.ArgumentParser(
        description="Apply transferred neuron IDs to manual_annotation.xlsx",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--target", required=True, help="Path to target recording")
    parser.add_argument("--matches", required=True,
                        help="Path to candidate_matches.csv from transfer_neuron_ids.py")
    parser.add_argument("--min_consensus", type=float, default=None,
                        help="Override minimum consensus fraction (0-1). "
                             "Default: use 'accepted' column from CSV.")
    parser.add_argument("--overwrite_existing", action="store_true",
                        help="Overwrite neurons that already have an ID1 value")
    parser.add_argument("--dry_run", action="store_true",
                        help="Show what would change without writing")
    parser.add_argument("--certainty", type=str, default="transfer",
                        help="Value to write in Certainty column (default: 'transfer')")
    args = parser.parse_args()

    anno_path = Path(args.target) / "3-tracking" / "manual_annotation" / "manual_annotation.xlsx"
    if not anno_path.exists():
        logger.error(f"manual_annotation.xlsx not found at {anno_path}")
        return

    # Load matches
    matches_path = Path(args.matches)
    if not matches_path.exists():
        logger.error(f"Matches file not found: {matches_path}")
        return

    with open(matches_path, "r") as f:
        reader = csv.DictReader(f)
        all_matches = list(reader)

    # Filter to accepted matches
    if args.min_consensus is not None:
        accepted = [
            m for m in all_matches
            if float(m["consensus_fraction"]) >= args.min_consensus
        ]
        logger.info(f"Filtered by consensus >= {args.min_consensus}: "
                     f"{len(accepted)}/{len(all_matches)} matches")
    else:
        accepted = [m for m in all_matches if m["accepted"] == "yes"]
        logger.info(f"Using 'accepted' column: {len(accepted)}/{len(all_matches)} matches")

    if not accepted:
        logger.warning("No matches to apply. Exiting.")
        return

    # Load existing annotations
    from openpyxl import load_workbook

    wb = load_workbook(str(anno_path))
    ws = wb.active

    # Build map: neuron_id -> row_number (1-indexed, data starts at row 2)
    neuron_rows = {}
    existing_ids = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        neuron_id = str(row[0].value).strip() if row[0].value else None
        if neuron_id:
            neuron_rows[neuron_id] = row_idx
            id1_val = row[2].value  # Column C = ID1
            if id1_val is not None and str(id1_val).strip():
                existing_ids[neuron_id] = str(id1_val).strip()

    logger.info(f"Existing annotations: {len(existing_ids)}/{len(neuron_rows)} neurons have ID1")

    # Determine what to write
    to_write = []
    skipped_existing = 0
    skipped_not_found = 0

    for m in accepted:
        target_nid = m["target_neuron_id"]
        predicted_id = m["predicted_id"]

        if target_nid not in neuron_rows:
            skipped_not_found += 1
            logger.debug(f"  {target_nid}: not found in xlsx, skipping")
            continue

        if target_nid in existing_ids and not args.overwrite_existing:
            skipped_existing += 1
            old_id = existing_ids[target_nid]
            match_str = "MATCH" if old_id == predicted_id else "MISMATCH"
            logger.debug(f"  {target_nid}: already has '{old_id}', predicted '{predicted_id}' ({match_str})")
            continue

        to_write.append({
            "target_neuron_id": target_nid,
            "predicted_id": predicted_id,
            "consensus_fraction": m["consensus_fraction"],
            "row_idx": neuron_rows[target_nid],
            "had_existing": target_nid in existing_ids,
            "old_id": existing_ids.get(target_nid, None),
        })

    # Summary
    print(f"\n{'='*70}")
    print(f"APPLY TRANSFER {'(DRY RUN)' if args.dry_run else ''}")
    print(f"{'='*70}")
    print(f"Target:              {args.target}")
    print(f"Accepted matches:    {len(accepted)}")
    print(f"To write:            {len(to_write)}")
    print(f"Skipped (existing):  {skipped_existing}")
    print(f"Skipped (not found): {skipped_not_found}")

    if to_write:
        print(f"\nChanges:")
        for w in to_write:
            if w["had_existing"]:
                print(f"  {w['target_neuron_id']}: '{w['old_id']}' -> '{w['predicted_id']}' "
                      f"(consensus: {w['consensus_fraction']})")
            else:
                print(f"  {w['target_neuron_id']}: (empty) -> '{w['predicted_id']}' "
                      f"(consensus: {w['consensus_fraction']})")

    # Accuracy check: how many overwritten values match vs mismatch?
    if args.overwrite_existing:
        n_match = sum(1 for w in to_write if w["had_existing"] and w["old_id"] == w["predicted_id"])
        n_mismatch = sum(1 for w in to_write if w["had_existing"] and w["old_id"] != w["predicted_id"])
        if n_match + n_mismatch > 0:
            print(f"\nOverwrite check: {n_match} match existing, {n_mismatch} differ")

    # Also check: among skipped_existing, how many would have been correct?
    if skipped_existing > 0:
        correct_skips = 0
        wrong_skips = 0
        for m in accepted:
            nid = m["target_neuron_id"]
            if nid in existing_ids and not args.overwrite_existing:
                if existing_ids[nid] == m["predicted_id"]:
                    correct_skips += 1
                else:
                    wrong_skips += 1
        print(f"\nSkipped neurons check: {correct_skips} predictions match existing, "
              f"{wrong_skips} differ (potential errors in existing or predicted)")

    print(f"{'='*70}")

    if args.dry_run:
        print("DRY RUN — no files modified.\n")
        wb.close()
        return

    if not to_write:
        print("Nothing to write.\n")
        wb.close()
        return

    # Backup
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = anno_path.parent / f"manual_annotation_backup_{timestamp}.xlsx"
    shutil.copy2(anno_path, backup_path)
    logger.info(f"Backup saved to {backup_path}")

    # Write changes
    for w in to_write:
        row_idx = w["row_idx"]
        ws.cell(row=row_idx, column=3, value=w["predicted_id"])  # Column C = ID1
        ws.cell(row=row_idx, column=5, value=args.certainty)     # Column E = Certainty

    wb.save(str(anno_path))
    wb.close()
    logger.info(f"Wrote {len(to_write)} IDs to {anno_path}")

    # Save apply log
    log_dir = Path(args.target) / "3-tracking" / "id_transfer"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"apply_log_{timestamp}.json"
    log_data = {
        "timestamp": timestamp,
        "matches_file": str(args.matches),
        "n_written": len(to_write),
        "n_skipped_existing": skipped_existing,
        "overwrite_existing": args.overwrite_existing,
        "min_consensus": args.min_consensus,
        "certainty_value": args.certainty,
        "backup_path": str(backup_path),
        "changes": [
            {
                "neuron_id": w["target_neuron_id"],
                "predicted_id": w["predicted_id"],
                "old_id": w["old_id"],
                "consensus": w["consensus_fraction"],
            }
            for w in to_write
        ],
    }
    with open(log_path, "w") as f:
        json.dump(log_data, f, indent=2)

    print(f"\nApply log saved to: {log_path}")
    print(f"Backup at: {backup_path}")
    print(f"Done.\n")


if __name__ == "__main__":
    main()
